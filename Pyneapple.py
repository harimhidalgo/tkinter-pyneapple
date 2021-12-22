# -*- coding: utf-8 -*-
"""
Created on Thu Dec 16 09:45:45 2021

@author: harim
"""

### IMPORTS AND DICTIONARIES ####################################################

# GUI IMPORTS
from tkinter import Tk, Frame, Button, Label, Menu, Toplevel, Y, Entry, Listbox, END, DISABLED, ACTIVE, IntVar, Scrollbar
from tkinter.ttk import Combobox

# PDF MERGER IMPORT
from PyPDF2 import PdfFileMerger

# MACRO RECORDER IMPORTS
from pynput.mouse import Listener as ML
from pynput.keyboard import Listener as KL, Key
from pyautogui import click, press, hotkey, scroll, keyDown, keyUp
from pyperclip import copy, paste
from time import sleep

# SMART LAUNCHER IMPORTS
from openpyxl import load_workbook
import webbrowser as wb

# OTHER IMPORTS
from os import listdir, path, remove, startfile

# Used to simply hit the key instead of pressing down and then up
single_keys = ['\t', '\n', '\r', ' ', '!', '"', '#', '$', '%', '&', "'", '(',
    ')', '*', '+', ',', '-', '.', '/', '0', '1', '2', '3', '4', '5', '6', '7',
    '8', '9', ':', ';', '<', '=', '>', '?', '@', '[', '\\', ']', '^', '_', '`',
    'a', 'b', 'c', 'd', 'e','f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o',
    'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '{', '|', '}', '~']

# Used to translate common hotkeys to the format used by pyautogui
translate_hotkey = {'\\x01' : ['ctrl', 'a'],
                    '\\x02' : ['ctrl', 'b'],
                    '\\x03' : ['ctrl', 'c'],
                    '\\x05' : ['ctrl', 'e'],
                    '\\x04' : ['ctrl', 'd'],
                    '\\x06' : ['ctrl', 'f'],
                    '\\x07' : ['ctrl', 'g'],
                    '\\x08' : ['ctrl', 'h'],
                    '\\x09' : ['ctrl', 'i'],
                    '\\x0a' : ['ctrl', 'j'],
                    '\\x0b' : ['ctrl', 'k'],
                    '\\x0c' : ['ctrl', 'l'],
                    '\\x0d' : ['ctrl', 'm'],
                    '\\x0e' : ['ctrl', 'n'],
                    '\\x0f' : ['ctrl', 'o'],
                    '\\x10' : ['ctrl', 'p'],
                    '\\x11' : ['ctrl', 'q'],
                    '\\x12' : ['ctrl', 'r'],
                    '\\x13' : ['ctrl', 's'],
                    '\\x14' : ['ctrl', 't'],
                    '\\x15' : ['ctrl', 'u'],
                    '\\x16' : ['ctrl', 'v'],
                    '\\x17' : ['ctrl', 'w'],
                    '\\x18' : ['ctrl', 'x'],
                    '\\x19' : ['ctrl', 'y'],
                    '\\x1a' : ['ctrl', 'z']}

# Used to translate special key names to the ones used by pyautogui
translate_key = {"ctrl_l":"ctrl",
                 "ctrl_r":"ctrl",
                 "ctrl":"ctrl",
                 "caps_lock":"capslock",
                 "alt":"alt",
                 "alt_l":"alt",
                 "alt_r":"altright",
                 "alt_gr":"altright",
                 "page_up":"pageup",
                 "page_down":"pagedown",
                 "cmd":"win",
                 "print_screen":"printscreen",
                 "media_previous":"prevtrack",
                 "media_play_pause":"playpause",
                 "media_next":"nexttrack",
                 "shift":"shift",
                 "shift_l":"shift",
                 "shift_r":"shift",
                 "media_volume_mute":"volumemute",
                 "media_volume_down":"volumedown",
                 "media_volume_up":"volumeup",
                 "\\\\":"\\",
                 "\"\"":"'"}

# Used to release all the keys before a fail-safe
additional_keys = ['accept', 'add', 'alt', 'altleft', 'altright', 'apps', 'backspace',
    'browserback', 'browserfavorites', 'browserforward', 'browserhome',
    'browserrefresh', 'browsersearch', 'browserstop', 'capslock', 'clear',
    'convert', 'ctrl', 'ctrlleft', 'ctrlright', 'decimal', 'del', 'delete',
    'divide', 'down', 'end', 'enter', 'esc', 'escape', 'execute', 'f1', 'f10',
    'f11', 'f12', 'f13', 'f14', 'f15', 'f16', 'f17', 'f18', 'f19', 'f2', 'f20',
    'f21', 'f22', 'f23', 'f24', 'f3', 'f4', 'f5', 'f6', 'f7', 'f8', 'f9',
    'final', 'fn', 'hanguel', 'hangul', 'hanja', 'help', 'home', 'insert', 'junja',
    'kana', 'kanji', 'launchapp1', 'launchapp2', 'launchmail',
    'launchmediaselect', 'left', 'modechange', 'multiply', 'nexttrack',
    'nonconvert', 'num0', 'num1', 'num2', 'num3', 'num4', 'num5', 'num6',
    'num7', 'num8', 'num9', 'numlock', 'pagedown', 'pageup', 'pause', 'pgdn',
    'pgup', 'playpause', 'prevtrack', 'print', 'printscreen', 'prntscrn',
    'prtsc', 'prtscr', 'return', 'right', 'scrolllock', 'select', 'separator',
    'shift', 'shiftleft', 'shiftright', 'sleep', 'space', 'stop', 'subtract', 'tab',
    'up', 'volumedown', 'volumemute', 'volumeup', 'win', 'winleft', 'winright', 'yen',
    'command', 'option', 'optionleft', 'optionright']

### FUNCTIONS ##################################################################

def read_config_file(name):
    ''' Parses the configurations in the .txt file and stores them in a dictionary. '''
    
    lines = []
    with open(name + '.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        file.close()
    lines = [ x.split(": ") for x in lines ]
    
    return dict(lines)

def macro_recorder():
    ''' Records all the keystrokes and mouse clicks, then executes them n number of times. '''

    fail_safe_log = []
    log = []      
        
    def on_click(x, y, button, pressed):
        ''' Logs the mouse clicks (position and button used). '''
        if pressed:
            log.append([x, y, str(button)[7:]])
    
    def on_scroll(x, y, dx, dy):
        ''' Logs the mouse scrolls (direction and distance). '''
        log.append([x, y, dx, dy])
        
    def on_press(key):
        ''' Logs the key name when pressed. '''
        if "Key." in str(key):        
            log.append([str(key)[4:].replace("'", ""), "down"])
        else:
            log.append([str(key).replace("'", ""), "down"]) 
        if key == Key.esc:
            return False
    
    def on_release(key):
        ''' Logs the key name when released. '''
        if "Key." in str(key):        
            log.append([str(key)[4:], "up"])            
        else:
            log.append([str(key).replace("'", ""), "up"]) 
    
    def fail_safe(key):
        ''' Stores the pressed keys in the fail_safe_log for later stop of the execution.'''
        fail_safe_log.append(key)
        if key == Key.esc:
            return False
        
    def record_instructions():    
        ''' Starts the listener after a given window of time. Captures all the mouse
            scrolls and clicks as wells as the keyboard strokes in a log.'''
        
        mouse_listener = ML(on_click=on_click, on_scroll=on_scroll)
        mouse_listener.start()
        with KL(on_press=on_press, on_release=on_release) as keyboard_listener:
            keyboard_listener.join()
            log.pop()
        mouse_listener.stop()
            
        return log
    
    def translate_instructions(instructions_list):
        ''' Takes an instructions list generated by the record_instructions() function \n
            and returns a list with friendlier names for the user. '''
        
        if instructions_list == []:
            return []
        
        result = []
        for i in instructions_list:
            if len(i) == 2:    
                if i[0] in single_keys and i[1] == "down": 
                    result.append("Hit {} key".format(i[0]))
                elif i[0] in translate_hotkey.keys() and i[1] == "down":
                    result.append("Send hotkey {} + {}".format(translate_hotkey[i[0]][0], translate_hotkey[i[0]][1]))
                elif i[0] in translate_key.keys():
                    if i[1] == "down":
                        result.append("Press down {} key".format(translate_key[i[0]]))
                    else:
                        result.append("Press up {} key".format(translate_key[i[0]]))
                else:
                    if i[1] == "down":                        
                        result.append("Hit {} key".format(i[0]))
            elif len(i) == 3: 
                result.append('Mouse clicked ({}, {}) with {} button'.format(i[0], i[1], i[2]))
            else:
                result.append('Mouse scrolled at ({}, {}) ({}, {})'.format(i[0], i[1], i[2], i[3]))
        
        return result
            
    
    def rec_str_to_list(rec_str):
        ''' Takes a string (that corresponds to the list generated by the
            record_instructions function) and turns it into a list again.'''        
        try:
            # for each instruction, split the info and store each its values
            rec_str = rec_str[2:-2].split('], [')
            results = []
            [ results.append(ins.split(', ')) for ins in rec_str ]
            for res in results:
                if len(res) == 2:   # remove the commas to keep the string as it is
                    results[results.index(res)] = [ x[1:-1] for x in res ]
                elif len(res) == 3: # takes [a, b, c] and returns [int(a), int(b), str(c[1:-1])]
                    results[results.index(res)][0] = int(res[0])
                    results[results.index(res)][1] = int(res[1])
                    results[results.index(res)][2] = str(res[2][1:-1])
                elif len(res) == 4: # converts each item of the list into an integer
                    results[results.index(res)] = [ int(x) for x in res ]
                else:
                    return []
            return results
        except:
            return []    

        
    def execution_screen(instructions=[]):
        ''' Displays the instructions recorded and the possible actions. '''
        
        def execute_instructions(instructions_list):
            ''' Takes a log with recorded instructions in pynput format and
                translates them into pyautogui format for execution.'''
            
            if instructions_list == []:
                return 
            
            # Start the listener to be able to fail-safe the executions with 'esc'
            KB_listener = KL(on_press=fail_safe)
            KB_listener.start()
            fail_safe_presses = []
            
            for i in instructions_list:
                sleep(0.08)    
                # If 'esc' key is pressed, stop the execution and exit
                if Key.esc in fail_safe_log:
                    KB_listener.stop()
                    if not fail_safe_presses == []: # realease all keys
                        for k in fail_safe_presses:
                            keyUp(k)
                    break
                
                if len(i) == 2: # check if the intruction is a key stroke
                    
                    if i[0] in single_keys and i[1] == "down": # check if the key is a character
                        if i[0] not in fail_safe_presses: fail_safe_presses.append(i[0])
                        press(i[0])
                    
                    elif i[0] in translate_hotkey.keys() and i[1] == "down": # check if it's a hotkey combination
                        hotkey(translate_hotkey[i[0]][0], translate_hotkey[i[0]][1])
                                        
                    elif i[0] in translate_key.keys(): # check if it's an special key
                        if i[1] == "down":
                            if translate_key[i[0]] not in fail_safe_presses: fail_safe_presses.append(translate_key[i[0]])
                            keyDown(translate_key[i[0]])                        
                        else:
                            keyUp(translate_key[i[0]])
                                            
                    else: # try-except block for the remaining keys
                        try:
                            if i[1] == "down":
                                if i[0] not in fail_safe_presses: fail_safe_presses.append(i[0])
                                press(i[0])                            
                        except:
                            pass
        
                elif len(i) == 3: # check if the intruction is a click  
                    click(i[0], i[1], button=i[2], duration=0.15)
                                
                else: # scroll the mouse
                    scroll(200*i[3], x=i[0], y=i[1])
                    
        def execution_main(secs, instructions_list, iterations):
            ''' Changes the execution screen display and executes the instructions. '''
            
            if secs == 0:
                btnx["state"] = DISABLED
                btny["state"] = DISABLED
                btnz["state"] = DISABLED
                entry_i["state"] = DISABLED
                main.update()
                try:
                    for i in range(iterations): 
                        label_1["text"] = str(i+1)+"/"+str(iterations)+" in progress..."
                        main.update()
                        execute_instructions(instructions_list)
                except:
                    pass
                if inst == []:
                    main.destroy()
                    return
                main.destroy()
            else:
                label_1["text"] = "Start in: " + str(secs) + " seconds"
                main.update()
        
        def execution_timer(timer, instructions_list, iterations):
            ''' Starts the execution after a given window of time. '''
            
            if timer == 0:
                main.after(1000, execution_main(timer, instructions_list, iterations))
                return
            else:
                main.after(1000, execution_main(timer, instructions_list, iterations))
                execution_timer(timer-1, instructions_list, iterations)
        
        main = Toplevel()
    
        main.geometry("480x280")
        main.geometry("+420+200")
        main.resizable(False, False)
        
        frame = Frame(main)
        frame.pack(fill=Y)
        
        mainmenu = Menu(frame)
        mainmenu.add_command(label="             Cancel             ", command=main.destroy)  
        main.config(menu=mainmenu)
        
        # INITIALIZE THE ACTIONS BUTTONS AND THE LABELS
        
        iters = IntVar()
        label_1 = Label(frame, text="Instructions:")
        label_2 = Label(frame, text="Iterations:")
        entry_i = Entry(frame, textvariable=iters, width=4)
        entry_i.insert(1, 1)
        scrollb = Scrollbar(frame)
        listbox = Listbox(frame, yscrollcommand=scrollb.set)  
        
        label_1.grid(row=0, column=0, pady=5, ipady=5)
        label_2.grid(row=0, column=1, pady=5, ipadx=5, sticky="E")
        entry_i.grid(row=0, column=2, pady=5, sticky="W")
        scrollb.grid(row=1, column=3, ipady=55, sticky="W")
        listbox.grid(row=1, column=0, columnspan=3, ipadx=100)
        scrollb.config(command=listbox.yview)
        
        # READ THE CONFIGURATIONS AND GET THE SECONDS FOR THE TIMER
        
        actual_configs = read_config_file("configurations")
        wait = int(actual_configs["rectime"])
    
        btnx = Button(frame, text= "Start execution \nright away", relief="groove",
                      command=lambda:[execution_timer(timer=0, instructions_list=instructions, iterations=iters.get())])
        btny = Button(frame, text= "Start execution \nin "+str(wait)+" seconds", relief="groove",
                      command=lambda:[execution_timer(timer=wait, instructions_list=instructions, iterations=iters.get())])
        btnz = Button(frame, text= "Save to \nclipboard", relief="groove",
                      command=lambda:[copy(str(instructions)), main.destroy()])
    
        btnx.grid(row=2, column=0, padx=5, pady=5, ipadx=10, ipady=5)
        btny.grid(row=2, column=1, padx=5, pady=5, ipadx=10, ipady=5)
        btnz.grid(row=2, column=2, columnspan=2, padx=5, pady=5, ipadx=10, ipady=5)

        # DISPLAY THE INSTRUCTIONS RECORDED
        
        i=1
        for ins in translate_instructions(instructions):    
            listbox.insert(i, "[{}] {}".format(str(i), ins))
            i+=1
                
        main.title("Macro Recorder")           
    
    def recording_main(secs):
        ''' Changes the screen display and starts recording the instructions. '''
        
        if secs == 0:
            label_1["text"] = "\n\n\n\n * * * NOW RECORDING * * * \n\n\n\n\n\n"
            btns["state"] = DISABLED
            btnt["state"] = DISABLED
            btnc["state"] = DISABLED
            btnc["text"] = "Reload to use\nthe clipboard"
            main.update()
            inst = record_instructions()
            if inst == []:
                label_1["text"] = "\n\n\n\nERROR: No instructions recorded.\n              Please try again.\n\n\n\n\n"
                btns["state"] = ACTIVE
                btnt["state"] = ACTIVE
                main.update()
                return
            main.destroy()
            execution_screen(inst)
        else:
            label_1["text"] = "\n\n\n\nStart recording in: " + str(secs) + " seconds\n\n\n\n\n\n"
            main.update()
    
    def recording_timer(timer=0):
        ''' Starts recording after a given window of time. '''
        
        if timer == 0:
            main.after(1000, recording_main(timer))
            return
        else:
            main.after(1000, recording_main(timer))
            recording_timer(timer-1)
    
    main = Toplevel()
    
    main.geometry("480x280")
    main.geometry("+420+200")
    main.resizable(False, False)
    
    frame = Frame(main)
    frame.pack(fill=Y)
    
    mainmenu = Menu(frame)
    mainmenu.add_command(label="             Back             ", command=main.destroy)  
    main.config(menu=mainmenu)
    
    # READ THE ACTUAL CONFIGURATIONS AND DISPLAY THE WARNING MESSAGE
    
    warning_msg = [
        "********************* IMPORTANT INFO: *********************\n\n",
        " (1) To stop recording, press the 'esc' key once.\n\n",
        " (2) To stop the execution, press 'esc' or quickly move the\n",
        "       mouse to a corner of the screen.\n\n",
        "  WARNING: The actions performed are not reversible. If \n",
        "                      anything goes wrong, press 'esc' repeteadly.\n\n",
        "***************************************************************"]
    
    label_1 = Label(frame, text=' '.join(warning_msg), justify="left")
    label_1.grid(row=0, column=0, columnspan=3, pady=5, ipady=5)
    
    actual_configs = read_config_file("configurations")
    wait = int(actual_configs["rectime"])

    # INITIALIZE ACTION BUTTONS  
    
    inst = rec_str_to_list(paste())
    
    btns = Button(frame, text= "Start recording \nright away", relief="groove", 
                  command=recording_timer)
    btnt = Button(frame, text= "Start recording \nin " + str(wait).strip() + " seconds", relief="groove", 
                  command=lambda:[recording_timer(wait)])
    btnc = Button(frame, text= "Use a recording \nfrom clipboard", relief="groove", 
                  command=lambda:[execution_screen(inst), main.destroy()])
    
    btns.grid(row=1, column=0, padx=5, pady=5, ipadx=10, ipady=5)
    btnt.grid(row=1, column=1, padx=5, pady=5, ipadx=10, ipady=5)
    btnc.grid(row=1, column=2, padx=5, pady=5, ipadx=10, ipady=5)
    
    if inst == []:
        btnc["text"] = "No recording \non clipboard"
        btnc["state"] = DISABLED
        main.update()
        
    main.title("Macro Recorder")   
    
    
def PDF_merger():
    ''' Merges a batch of PDF files (*.pdf) into a single one. '''
    
    def list_pdfs_in(directory):
        ''' Takes a directory and return a list with the PDF files found. '''
        
        result = []
        directory = directory.strip()
        for file in listdir(directory):
            if file.endswith(".pdf"):
                result.append(path.basename(path.join(directory, file)))
        return result 
    
    def move_up(Listbox):
            ''' Moves the selected item of the listbox one place up. '''
            try:
                Listbox.idxs = Listbox.curselection()
                if not Listbox.idxs:
                    return
                for pos in Listbox.idxs:
                    if pos == 0:
                        continue
                    text=Listbox.get(pos)
                    Listbox.delete(pos)
                    Listbox.insert(pos-1, text)
                    Listbox.pop(pos)
                    Listbox.insert(pos-1, text)
                    Listbox.selection_set(pos)
            except:
                pass
            
    def move_down(Listbox, limit):
        ''' Moves the selected item of the listbox one place down. '''
        try:
            Listbox.idxs = Listbox.curselection()
            if not Listbox.idxs:
                return
            for pos in Listbox.idxs:
                if pos == limit: 
                    continue
                text=Listbox.get(pos)
                Listbox.delete(pos)
                Listbox.insert(pos+1, text)
                Listbox.pop(pos)
                Listbox.insert(pos+1, text)
                Listbox.selection_set(pos)
        except:
            pass
        
    def remove_item(Listbox):
        ''' Removes the selected item of the listbox if there's' 2+ items. '''
        
        try:
            Listbox.idxs = Listbox.curselection()
            if not Listbox.idxs or len(Listbox.get(0 , END)) <= 2:
                return
            Listbox.delete(Listbox.idxs)
        except:
            pass
        
    def btn_functionality(Button, Listbox, function, path_to_pdfs="", pdf_name="", delete=""):
        ''' Deteles the current item or merges all items in the listbox, if possible. \n
            When function="merge", optional argumets must be passed to the function. '''
        
        if function == "merge":
            if len(Listbox.get(0 , END)) < 2: # Then is impossible to complete the merging
                Button["state"] = DISABLED
                return
            else:
                try:  
                    merging_list = Listbox.get(0 , END)
                    merger = PdfFileMerger()   
                    for pdf_file in merging_list:
                        merger.append(path.join(path_to_pdfs.strip(), pdf_file))                
                    merged_pdf_path = path.join(path_to_pdfs.strip(), pdf_name.strip()+'.pdf')
                    merger.write(merged_pdf_path)
                    merger.close()
                    startfile(merged_pdf_path)
                    if delete.strip() == "True (not recommended)":
                        for pdf_file in merging_list:
                            file_to_remove = path.join(path_to_pdfs.strip(), pdf_file)
                            if not file_to_remove == path.join(path_to_pdfs.strip(), pdf_name.strip()+'.pdf'):
                                remove(file_to_remove)               
                    else:
                        return
                except:
                    pass
        elif function == "remove":
            if len(Listbox.get(0 , END)) <= 2: # Then is not logic to remove any more items
                Button["state"] = DISABLED
                return
            else:
                remove_item(Listbox)
        else:
                pass       

    
    main = Toplevel()
    
    main.geometry("480x280")
    main.geometry("+420+200")
    main.resizable(False, False)
    
    frame = Frame(main)
    frame.pack(fill=Y)
    
    mainmenu = Menu(frame)
    mainmenu.add_command(label="             Back             ", command=main.destroy)  
    main.config(menu=mainmenu)
    
    # READ THE ACTUAL CONFIGURATIONS TO LIST THE PDFS
    
    scrollb = Scrollbar(frame)
    label_1 = Label(frame, text="Files selected for merging:")
    listbox = Listbox(frame, yscrollcommand=scrollb.set)  
    
    actual_configs = read_config_file("configurations")
    merging_list = list_pdfs_in(actual_configs["pdfpath"])
    
    if merging_list == []:
        listbox.insert(1, "ERROR: There are no PDF files to merge.")
        listbox.insert(2, "Source directory 🡆 " + actual_configs["pdfpath"])
        label_1["text"] = "Check your folder/configurations."
    elif len(merging_list) == 1:
        listbox.insert(1, "ERROR: You need more than one PDF file.")
        listbox.insert(2, "File found: " + actual_configs["pdfname"] + ".pdf")        
        label_1["text"] = "Check your folder/configurations."
    else:
        i=1
        for file in merging_list:    
            listbox.insert(i, file)
            i+=1

        # ENABLE THE MERGING BUTTON

        btnm = Button(frame, text= "Start merging", relief="groove", 
                  command=lambda:[btn_functionality(btnm, listbox, "merge", 
                                                    path_to_pdfs=actual_configs["pdfpath"], 
                                                    pdf_name=actual_configs["pdfname"], 
                                                    delete=actual_configs["pdfkeep"]), 
                                  main.destroy()])    
        btnm.grid(row=0, column=2, pady=5, ipadx=5, ipady=5, sticky="E")

    # INITIALIZE ACTION BUTTONS  
    
    scrollb.grid(row=1, column=3, ipady=60, sticky="W")
    label_1.grid(row=0, column=0, columnspan=2, pady=5, ipady=5, sticky="W")
    listbox.grid(row=1, column=0, columnspan=3, ipadx=100, ipady=5)
    scrollb.config(command=listbox.yview)
        
    btnu = Button(frame, text= "▲", relief="groove", command=lambda:[move_up(listbox)])
    btnd = Button(frame, text= "▼", relief="groove", command=lambda:[move_down(listbox, listbox.curselection())])
    btnr = Button(frame, text= "Remove", relief="groove", command=lambda:[btn_functionality(btnr, listbox, "remove")])  
    
    btnu.grid(row=2, column=0, padx=5, pady=5, ipadx=10, ipady=5, sticky="E")
    btnd.grid(row=2, column=1, padx=5, pady=5, ipadx=10, ipady=5, sticky="W")
    btnr.grid(row=2, column=2, padx=5, pady=5, ipadx=10, ipady=5)
    
    main.title("PDF Merger")   
    

def smart_launcher():
    ''' Allows quicker access to folders, files and executables. '''
    
    def read_xlms(file_path):
        ''' Loads the Excel file (from file_path) and returns a dictionary with \n
            the value of each column in range A1:T20. '''
                
        try: wb = load_workbook(file_path.strip())        
        except: return {}
        sheets = wb.sheetnames
        sheet = wb[sheets[0]]
        
        result = {}        
        for column in range(1,21):
            cv = sheet.cell(1, column).value
            if cv != None:
                result[cv] = []
                
                for row in range(1,21):
                    rv = sheet.cell(row, column).value
                    if rv != None and not row == 1:                
                        if rv[:4].lower() == "open" or rv[:6].lower() == "browse":
                            result[cv].append(rv)
        
        return result
    
    def launch_actions(actions_list):
        ''' Parses the actions from the list received and executes them. '''
        
        if actions_list == []:
            return
        
        for action in actions_list:
            try:
                if action[:4].lower() == "open":
                    startfile(action[5:])
                elif action[:6].lower() == "browse":
                    wb.open(action[7:])
            except:
                pass
            
        main.destroy()
    
    def search_btn(search_in, key_word, Combobox, Entry, Button):
        ''' Changes the display of the screen according to the matches found, \n
            updating the Combobox's selection, the Entry's text and the Button's label. '''
        
        next_index = 0
        matches = []
        
        if len(key_word) >= 1:
            for key in search_in:
                if key_word in key:
                    matches.append(key)
                    if Combobox.get() == key:
                        next_index = matches.index(key) + 1 
        
            if len(matches) == 1 and next_index == 0: # unique coincidence
                Combobox.set(matches[0])
                Button["text"] = "Search"
                Entry.delete(0, END)
            elif len(matches) == next_index and len(matches) > 0: # is the last element of the search
                Combobox.set(matches[0])
                Button["text"] = "Search"
                Entry.delete(0, END)
            elif len(matches) == 0: # no coincidence
                Button["text"] = "Search"
                Entry.delete(0, END)
            else: # more than one coincidence
                Combobox.set(matches[next_index])
                Button["text"] = "Search next"
        else:
            Button["text"] = "Search"
            
        return
    
    main = Toplevel()
    
    main.geometry("480x280")
    main.geometry("+420+200")
    main.resizable(False, False)
    
    frame = Frame(main)
    frame.pack(fill=Y)
    
    mainmenu = Menu(frame)
    mainmenu.add_command(label="             Back             ", command=main.destroy)  
    main.config(menu=mainmenu)
    
    # READ THE ACTUAL CONFIGURATIONS AND THE EXCEL FILE    
    
    actual_configs = read_config_file("configurations")
    actions = read_xlms(actual_configs["laupath"])   
    optlist = list(actions.keys())
    
    # INITIALIZE ALL THE WIDGETS
    
    label_1 = Label(frame, text="Search by keyword:")
    label_2 = Label(frame, text="Select the actions set:")
    entry_k = Entry(frame, width=30)
    actmenu = Combobox(frame, state="readonly", width=30)
    launchb = Button(frame, text= "Launch actions set", relief="groove", 
                     command=lambda:[launch_actions(actions[actmenu.get()])])
    searchb = Button(frame, text= "Search", relief="groove", 
                     command=lambda:[search_btn(optlist, entry_k.get(), actmenu, entry_k, searchb)])
                                     
    label_2.grid(row=0, column=0, padx=10, pady=25, ipady=5, sticky="E")
    actmenu.grid(row=0, column=1, pady=25, ipady=2, sticky="W")    
    label_1.grid(row=1, column=0, padx=10, pady=25, ipady=5, sticky="E")
    entry_k.grid(row=1, column=1, pady=25, ipady=2, sticky="W")
    launchb.grid(row=2, column=1, padx=5, pady=25, ipadx=15, ipady=10, sticky="E")
    searchb.grid(row=2, column=0, padx=5, pady=25, ipadx=15, ipady=10, sticky="E")
    
    if actions == {}:
        label_1["text"] = "ERROR: Unable get actions \n from the Excel file."
        label_2["text"] = "Path to the Excel file 🡆 " 
        entry_k.insert(0, "Check your file/configurations.")
        entry_k["state"] = DISABLED
        launchb["state"] = DISABLED
        searchb["state"] = DISABLED
        actmenu.set(actual_configs["laupath"].strip())        
    else:
        actmenu["values"] = optlist
        actmenu.set(optlist[0])
    
    main.title("Smart Launcher")   
    

### MENU FUNCTIONS ##############################################################

def open_about():
    ''' Opens a new window with additional information about the program. '''
    
    main = Toplevel()
    
    main.geometry("480x280")
    main.geometry("+420+200")
    
    frame = Frame(main)
    frame.pack(fill=Y)
    
    mainmenu = Menu(frame)
    mainmenu.add_command(label="             Back             ", command=main.destroy)  
    main.config(menu=mainmenu)
    
    lb1 = Label(frame, text="@HarimHidal on GitHub")
    lb1.grid(row=0, column=0, ipadx=25, ipady=25)
    
    main.title("About")    
            
def open_config():
    ''' Opens the configuration window to change the content of the configurations.txt file. '''
    
    def save_configs(new_configs):
        ''' Saves the settings to configurations.txt by updating the default_configs dictionary with the one received. '''
    
        default_configs = read_config_file("default")
        for k,v in new_configs.items():
            new_configs[k] = v.replace("\n","") + '\n'
        default_configs.update(new_configs)
        
        new_configs_list = []
        [ new_configs_list.append(str(k) + ": " + str(v)) for k, v in default_configs.items() ]
        
        with open('configurations.txt', 'w', encoding='utf-8') as file:
            file.writelines(new_configs_list)
            file.close()
            
    main = Toplevel()
    
    main.geometry("480x300")
    main.geometry("+420+200")
    main.resizable(False, False)
    
    frame = Frame(main)
    frame.pack(fill=Y)
    
    # READ THE ACTUAL CONFIGURATIONS
    
    actual_configs = read_config_file("configurations")    
    
    # MACRO RECORDER CONFIGURATIONS
    
    rec0 = Label(frame, text="Macro \nRecorder:", font="Heveltica 10")
    rec1 = Label(frame, text="Seconds for the timer:")
    
    rec0.grid(row=0, column=0, padx=5, pady=15)
    rec1.grid(row=0, column=1, padx=5, pady=5, sticky="E")
    
    vlist2 = ["3", "5", "10", "30", "60"]
    rectime = Combobox(frame, state="readonly", values=vlist2,  width=3)
    rectime.set(str(actual_configs["rectime"]))
    
    rectime.grid(row=0, column=2, padx=5, pady=5, sticky="W")
    
    # PDF MERGER CONFIGURATIONS
    
    pdf0 = Label(frame, text="PDF \nMerger:", font="Heveltica 10")
    pdf1 = Label(frame, text="Path to the PDF files:")
    pdf2 = Label(frame, text="Merged file name:")
    pdf3 = Label(frame, text="Delete source files:")
    
    pdf0.grid(row=1, column=0, rowspan=4, padx=5, pady=5, sticky="N")
    pdf1.grid(row=2, column=1, padx=5, pady=5, sticky="E")
    pdf2.grid(row=3, column=1, padx=5, pady=5, sticky="E")
    pdf3.grid(row=4, column=1, padx=5, pady=5, sticky="E") 
    
    pdfpath = Entry(frame, width=35)
    pdfpath.insert(0, actual_configs["pdfpath"])
    
    pdfname = Entry(frame, width=35)
    pdfname.insert(0, actual_configs["pdfname"])
    
    vlist = ["False (recommended)", "True (not recommended)"]
    pdfkeep = Combobox(frame, state="readonly", values=vlist, width=22)
    pdfkeep.set(actual_configs["pdfkeep"])
    
    pdfpath.grid(row=2, column=2, padx=5, pady=5, sticky="W")
    pdfname.grid(row=3, column=2, padx=5, pady=5, sticky="W")
    pdfkeep.grid(row=4, column=2, padx=5, ipadx=5, sticky="W") 
    
    # SMART LAUNCHER CONFIGURATIONS
    
    lau0 = Label(frame, text="Smart \nLauncher:", font="Heveltica 10")
    lau1 = Label(frame, text="Excel file location:")
    
    lau0.grid(row=5, column=0, padx=5, pady=15)
    lau1.grid(row=5, column=1, padx=5, pady=5, sticky="E")
    
    laupath = Entry(frame, width=35)
    laupath.insert(0, actual_configs["laupath"])
    
    laupath.grid(row=5, column=2, padx=5, pady=5, sticky="W")
    
    # INITIALIZE ACTION BUTTONS

    btn_cancel = Button(frame, text= "Cancel", relief="groove", command=main.destroy)
    btndefault = Button(frame, text= "Set default", relief="groove", 
                        command=lambda:[save_configs({}), main.destroy()])
    btnsaveall = Button(frame, text= "Save new configurations", relief="groove", 
                        command=lambda:[save_configs({"rectime": rectime.get(), 
                                                      "pdfpath": pdfpath.get(), 
                                                      "pdfname": pdfname.get(),
                                                      "pdfkeep": pdfkeep.get(), 
                                                      "laupath": laupath.get()}), main.destroy()])
    
    btn_cancel.grid(row=10, column=0, padx=10, pady=5, ipadx=10, ipady=7, sticky="W")
    btndefault.grid(row=10, column=1, padx=10, pady=5, ipadx=10, ipady=7, sticky="E")
    btnsaveall.grid(row=10, column=2, padx=10, pady=5, ipadx=10, ipady=7, sticky="E")
    
    main.title("Configurations")

### MAIN ########################################################################

main = Tk()

main.geometry("480x300")
main.geometry("+420+200")
main.resizable(False, False)

frame = Frame(main)
frame.pack(fill=Y)

# UPPER MENU BUTTONS

btna = Button(frame, text= "About", relief="flat", command=open_about)
btnc = Button(frame, text= "Configurations", relief="flat", command=open_config)
btne = Button(frame, text= "Exit", relief="flat", command=main.destroy)

btna.grid(row=0, column=0, ipadx=25, ipady=5, sticky="NSEW")
btnc.grid(row=0, column=1, ipadx=25, ipady=5, sticky="NSEW")
btne.grid(row=0, column=2, ipadx=25, ipady=5, sticky="NSEW")

# ACTION BUTTONS THAT CALL THE MAIN FUNCTIONS

lb0 = Label(frame, text="Select one of the options below:")  
btn1 = Button(frame, text= "Macro \n Recorder", font="Heveltica 11", relief="groove", command=macro_recorder)
btn2 = Button(frame, text="    PDF    \nMerger", font="Heveltica 11", relief="groove", command=PDF_merger)
btn3 = Button(frame, text="Smart \nLauncher", font="Heveltica 11", relief="groove", command=smart_launcher)

lb0.grid(row=1, column=0, columnspan=3, ipadx=25, ipady=25)
btn1.grid(row=2, column=0, ipadx=10, ipady=10, padx=20, pady=5)
btn2.grid(row=2, column=1, ipadx=10, ipady=10, padx=20, pady=5)
btn3.grid(row=2, column=2, ipadx=10, ipady=10, padx=20, pady=5)

# DESCRIPTIVE LABELS 

lb1 = Label(frame, text="Records all the \nkeystrokes and \nmouse clicks.")
lb2 = Label(frame, text="Merges a batch of \nfiles (*.pdf) into a \nsingle PDF file.")
lb3 = Label(frame, text="Allows quicker \naccess to folders, \nfiles and apps.")

lb1.grid(row=3, column=0, ipadx=5, ipady=5)
lb2.grid(row=3, column=1, ipadx=5, ipady=5)
lb3.grid(row=3, column=2, ipadx=5, ipady=5)

main.title("Pyneapple")

main.mainloop()


